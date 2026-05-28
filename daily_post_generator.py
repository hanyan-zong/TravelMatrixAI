"""BWS 小红书每日内容生成器

核心流程：
1. 读取 Excel「统一资源库」+ 辅助 Sheet（主题库/选题模板/图片Prompt/文案模板）
2. 按每日选题模板匹配今日主题 → 自动筛选主资源 + 搭配资源
3. 填充文案模板变量 → 生成小红书正文
4. 填充图片 Prompt 模板变量 → 输出占位 Prompt 文本
5. 生成 HTML 预览页（小红书竖版卡片样式）
6. 输出到 output/YYYY-MM-DD/ 目录

用法:
    python daily_post_generator.py                      # 生成今日内容
    python daily_post_generator.py 2026-05-27           # 生成指定日期内容
    python daily_post_generator.py --all                # 生成全部排期内容
    python daily_post_generator.py --date 2026-05-27    # 同上，显式参数
"""

from __future__ import annotations

import argparse
import json
import random
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

# ─── 路径常量 ────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
EXCEL_PATH = BASE_DIR / "BWS碎片化整理 - 2026年05月25日 chatGPT清洗过的.xlsx"
OUTPUT_DIR = BASE_DIR / "output"


# ═══════════════════════════════════════════════════════════════════════════════
# 1. 数据读取层
# ═══════════════════════════════════════════════════════════════════════════════

def _sheet_to_dicts(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    """把 openpyxl worksheet 转成 list[dict]，第一行为 header"""
    headers: list[str | None] = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row: dict[str, Any] = {}
        for c, h in enumerate(headers, 1):
            if h is None:
                continue
            row[h] = ws.cell(r, c).value
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def load_all_data(excel_path: Path = EXCEL_PATH) -> dict:
    """读取 Excel 所有关键 Sheet，返回结构化字典"""
    wb = openpyxl.load_workbook(str(excel_path), read_only=False, data_only=True)

    data = {
        "resources": _sheet_to_dicts(wb["统一资源库"]),
        "themes": _sheet_to_dicts(wb["内容主题库"]),
        "schedule": _sheet_to_dicts(wb["每日选题模板"]),
        "image_prompts": _sheet_to_dicts(wb["图片Prompt模板"]),
        "copy_templates": _sheet_to_dicts(wb["小红书文案模板"]),
    }

    wb.close()
    return data


# ═══════════════════════════════════════════════════════════════════════════════
# 2. 资源筛选层 — 根据主题自动匹配资源
# ═══════════════════════════════════════════════════════════════════════════════

# 主题 ID → 统一资源库中 resource_type 的映射
THEME_RESOURCE_TYPE_MAP = {
    "T001": "一日游/玩乐",
    "T002": "酒店/常规",
    "T003": "酒店/高端",       # 也可搭配 高端下午茶
    "T004": "小众路线/泗水",
    "T005": "科莫多/出海",
    "T006": "疗愈/瑜伽",
    "T007": "交通/车费",
}

# 主题 ID → 搭配资源类型（用于选 Mix1 / Mix2）
THEME_MIX_TYPE_MAP = {
    "T001": ["交通/车费", "酒店/常规"],
    "T002": ["交通/车费", "疗愈/瑜伽"],
    "T003": ["高端下午茶", "疗愈/瑜伽"],
    "T004": ["交通/车费", "酒店/常规"],
    "T005": ["酒店/常规", "交通/车费"],
    "T006": ["酒店/常规", "高端下午茶"],
    "T007": ["一日游/玩乐", "酒店/常规"],
}


def _parse_date(val) -> date | None:
    """尝试把 Excel 中的日期值转成 date 对象"""
    if isinstance(val, date):
        return val
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def get_schedule_for_date(schedule: list[dict], target_date: date) -> dict | None:
    """从每日选题模板中找到目标日期的选题行"""
    for row in schedule:
        row_date = _parse_date(row.get("发布日期"))
        if row_date == target_date:
            return row
    return None


def find_resource_by_id(resources: list[dict], rid: str) -> dict | None:
    """按 resource_id 查找资源"""
    for r in resources:
        if r.get("resource_id") == rid:
            return r
    return None


# 主题 ID → 用于匹配主资源的关键词（从主题名称提取核心地名/体验词）
THEME_KEYWORD_HINTS = {
    "T001": ["佩尼达", "蓝梦", "海岛", "出海", "浮潜"],
    "T002": ["酒店", "住宿"],
    "T003": ["高端", "蜜月", "度假", "下午茶"],
    "T004": ["泗水", "布罗莫", "伊真", "火山"],
    "T005": ["科莫多", "粉沙滩", "出海"],
    "T006": ["疗愈", "瑜伽", "乌布", "苍古", "SPA"],
    "T007": ["包车", "接送机", "交通", "司机"],
}


def _match_score(resource: dict, keywords: list[str]) -> int:
    """计算资源与关键词的匹配分数"""
    text = " ".join([
        _safe_str(resource.get("product_name_cn", "")),
        _safe_str(resource.get("visual_keywords", "")),
        _safe_str(resource.get("xhs_angle", "")),
        _safe_str(resource.get("area", "")),
        _safe_str(resource.get("destination", "")),
    ])
    score = 0
    for kw in keywords:
        if kw in text:
            score += 1
    return score


def pick_main_resource(resources: list[dict], theme_id: str, hint_id: str | None = None) -> dict:
    """根据主题选取主资源。如果 hint_id 有值则优先用，否则按资源类型+关键词匹配"""
    if hint_id:
        found = find_resource_by_id(resources, hint_id)
        if found:
            return found

    res_type = THEME_RESOURCE_TYPE_MAP.get(theme_id, "")
    candidates = [r for r in resources if r.get("resource_type") == res_type]
    if not candidates:
        # 退化：随机选一个
        candidates = resources

    # 按关键词匹配打分，优先选最匹配的（同分则随机）
    keywords = THEME_KEYWORD_HINTS.get(theme_id, [])
    if keywords:
        scored = [(r, _match_score(r, keywords)) for r in candidates]
        max_score = max(s for _, s in scored)
        top_candidates = [r for r, s in scored if s == max_score]
        return random.choice(top_candidates)

    return random.choice(candidates)


def pick_mix_resource(resources: list[dict], theme_id: str, main: dict, slot: int = 1) -> dict | None:
    """根据主题选取搭配资源（Mix1 / Mix2），避免与主资源重复，优先同区域"""
    mix_types = THEME_MIX_TYPE_MAP.get(theme_id, [])
    if not mix_types:
        return None

    # slot 1 对应 mix_types[0], slot 2 对应 mix_types[1]
    idx = slot - 1
    if idx >= len(mix_types):
        idx = 0
    target_type = mix_types[idx]

    main_area = _safe_str(main.get("area", ""))
    main_dest = _safe_str(main.get("destination", ""))

    candidates = [r for r in resources if r.get("resource_type") == target_type and r.get("resource_id") != main.get("resource_id")]
    if not candidates:
        return None

    # 优先同区域
    if main_area:
        same_area = [r for r in candidates if _safe_str(r.get("area", "")) == main_area]
        if same_area:
            return random.choice(same_area)

    # 优先同目的地
    if main_dest:
        same_dest = [r for r in candidates if _safe_str(r.get("destination", "")) == main_dest]
        if same_dest:
            return random.choice(same_dest)

    return random.choice(candidates)


# ═══════════════════════════════════════════════════════════════════════════════
# 3. 文案生成层 — 填充文案模板变量
# ═══════════════════════════════════════════════════════════════════════════════

def _safe_str(val) -> str:
    """安全转字符串，None → 空串"""
    if val is None:
        return ""
    return str(val).strip()


def build_copy_context(theme: dict, main: dict, mix1: dict | None, mix2: dict | None) -> dict:
    """构建文案模板变量上下文"""
    destination = _safe_str(main.get("destination", "巴厘岛"))
    area = _safe_str(main.get("area", ""))
    visual_keywords = _safe_str(main.get("visual_keywords", ""))
    xhs_angle = _safe_str(main.get("xhs_angle", ""))
    highlights = _safe_str(main.get("highlights", ""))
    price_text = _safe_str(main.get("price_text", ""))
    product_name = _safe_str(main.get("product_name_cn", ""))
    includes = _safe_str(main.get("includes", ""))
    excludes = _safe_str(main.get("excludes", ""))
    suitable_for = _safe_str(main.get("suitable_for", ""))
    caution = _safe_str(main.get("caution", ""))
    theme_name = _safe_str(theme.get("主题名称", ""))
    theme_angle = _safe_str(theme.get("选题角度", ""))
    cover_selling = _safe_str(theme.get("封面卖点", ""))
    # 目标人群：优先用主题库的「目标人群」字段，更精准
    target_crowd = _safe_str(theme.get("目标人群", ""))
    # suitable_for 字段目前是标签式内容，作为 fallback
    suitable_for = target_crowd or _safe_str(main.get("suitable_for", ""))

    # 核心体验：从 visual_keywords 中提取与旅行体验相关的关键词
    # 过滤掉"机场接送、舒适商务车、中文司机、行李"等工具性关键词
    _skip_keywords = {"机场接送", "舒适商务车", "中文司机", "行李", "交通便利", "酒店住宿"}
    raw_kws = [k.strip() for k in visual_keywords.replace("、", ",").split(",") if k.strip()]
    experience_kws = [k for k in raw_kws if k not in _skip_keywords]

    if experience_kws:
        core_experience = "、".join(experience_kws[:4])
    elif xhs_angle:
        # fallback: 用 xhs_angle，但去掉"第一次去巴厘岛"之类的通用前缀
        core_experience = xhs_angle.replace("第一次去", "").replace("必玩", "")
    else:
        core_experience = destination + "旅行"

    # 搭配资源信息
    mix1_info = ""
    if mix1:
        mix1_info = f"搭配推荐：{_safe_str(mix1.get('product_name_cn'))}（{_safe_str(mix1.get('price_text'))}）"

    mix2_info = ""
    if mix2:
        mix2_info = f"搭配推荐：{_safe_str(mix2.get('product_name_cn'))}（{_safe_str(mix2.get('price_text'))}）"

    # 同时提供中英文变量名，兼容不同模板的占位符格式
    # 中文变量名：{目的地}、{主题名称}、{核心体验} 等
    # 英文变量名：{destination}、{area}、{visual_keywords} 等
    return {
        # 中文变量（文案模板用）
        "目的地": destination,
        "区域": area,
        "主题": theme_name,
        "主题名称": theme_name,
        "visual_keywords": visual_keywords,
        "xhs_angle": xhs_angle,
        "核心体验": core_experience,
        "亮点": highlights,
        "价格": price_text,
        "产品名": product_name,
        "包含项": includes,
        "不含项": excludes,
        "适合人群": suitable_for,
        "注意事项": caution,
        "选题角度": theme_angle,
        "封面卖点": cover_selling,
        "搭配1": mix1_info,
        "搭配2": mix2_info,
        # 英文变量（图片 Prompt 模板用）
        "destination": destination,
        "area": area,
    }


def fill_copy_templates(copy_templates: list[dict], ctx: dict) -> dict[str, str]:
    """填充小红书文案 6 个模块，返回 {模块名: 填充后文本}"""
    filled: dict[str, str] = {}

    for tpl_row in copy_templates:
        module = _safe_str(tpl_row.get("模块"))
        content = _safe_str(tpl_row.get("模板内容"))
        if not module or not content:
            continue

        # 替换 {变量名} 格式的占位符
        result = content
        for key, val in ctx.items():
            result = result.replace(f"{{{key}}}", val)

        filled[module] = result

    return filled


def generate_full_post(filled_modules: dict, ctx: dict, theme: dict) -> str:
    """把填充好的文案模块组装成完整小红书帖子

    按照「正文结构」模板的 ①~⑥ 顺序，结合实际资源数据，生成结构化正文
    """
    parts = []

    destination = _safe_str(ctx.get("目的地", "巴厘岛"))
    product_name = _safe_str(ctx.get("产品名", ""))
    highlights = _safe_str(ctx.get("亮点", ""))
    price = _safe_str(ctx.get("价格", ""))
    includes = _safe_str(ctx.get("包含项", ""))
    excludes = _safe_str(ctx.get("不含项", ""))
    suitable = _safe_str(ctx.get("适合人群", ""))
    caution = _safe_str(ctx.get("注意事项", ""))
    core_experience = _safe_str(ctx.get("核心体验", ""))
    cover_selling = _safe_str(ctx.get("封面卖点", ""))
    mix1 = _safe_str(ctx.get("搭配1", ""))
    mix2 = _safe_str(ctx.get("搭配2", ""))

    # 标题
    title = filled_modules.get("标题公式", "")
    if title:
        parts.append(title)

    # 开头钩子
    hook = filled_modules.get("开头钩子", "")
    if hook:
        parts.append(f"\n{hook}")

    # 正文 — 按结构模板的 6 个模块来组织
    parts.append(f"\n① 为什么值得去")
    if highlights:
        parts.append(highlights)
    if cover_selling:
        parts.append(cover_selling)

    parts.append(f"\n② 行程安排")
    if product_name:
        parts.append(f"📍 {product_name}")
    if includes:
        parts.append(f"📦 一价全含：{includes}")

    parts.append(f"\n③ 适合谁")
    if suitable:
        parts.append(f"👥 {suitable}")
    else:
        parts.append(f"第一次去{destination}、想省心省力的小伙伴")

    parts.append(f"\n④ 费用")
    if price:
        parts.append(f"💰 {price}")
    if excludes:
        parts.append(f"❌ 不含：{excludes}")

    if caution:
        parts.append(f"\n⑤ 注意事项")
        parts.append(f"⚠️ {caution[:200]}{'...' if len(caution) > 200 else ''}")

    # 搭配推荐
    if mix1 or mix2:
        parts.append(f"\n⑥ 搭配推荐")
        if mix1:
            parts.append(f"🔗 {mix1}")
        if mix2:
            parts.append(f"🔗 {mix2}")

    # 卖点句
    selling = filled_modules.get("卖点句", "")
    if selling:
        parts.append(f"\n💡 {selling}")

    # CTA
    cta = filled_modules.get("转化CTA", "")
    if cta:
        parts.append(f"\n📩 {cta}")

    # 标签
    tags = filled_modules.get("标签池", "")
    if tags:
        parts.append(f"\n{tags}")

    return "\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════════
# 4. 图片 Prompt 生成层
# ═══════════════════════════════════════════════════════════════════════════════

# 主题 ID → 推荐使用的 prompt_type
THEME_PROMPT_TYPE_MAP = {
    "T001": "封面图",     # 海岛
    "T002": "酒店图",
    "T003": "酒店图",
    "T004": "火山图",
    "T005": "封面图",     # 海岛/出海
    "T006": "疗愈图",
    "T007": "轮播图",     # 路线攻略
}


def fill_image_prompts(image_prompts: list[dict], ctx: dict, theme_id: str) -> list[dict]:
    """根据主题匹配图片 Prompt 模板，填充变量，返回 [{prompt_type, filled_prompt, negative, aspect}]"""
    results = []

    for ip in image_prompts:
        prompt_type = _safe_str(ip.get("prompt_type"))
        template = _safe_str(ip.get("GPT-image-2 Prompt 模板"))
        negative = _safe_str(ip.get("负面/限制"))
        aspect = _safe_str(ip.get("画幅建议"))

        if not template:
            continue

        filled = template
        for key, val in ctx.items():
            filled = filled.replace(f"{{{key}}}", val)

        results.append({
            "prompt_type": prompt_type,
            "filled_prompt": filled,
            "negative": negative,
            "aspect": aspect,
        })

    return results


# ═══════════════════════════════════════════════════════════════════════════════
# 5. HTML 预览页生成
# ═══════════════════════════════════════════════════════════════════════════════

def generate_html_preview(
    target_date: date,
    schedule_row: dict,
    theme: dict,
    main: dict,
    mix1: dict | None,
    mix2: dict | None,
    post_text: str,
    filled_modules: dict,
    image_prompts_filled: list[dict],
    ctx: dict,
) -> str:
    """生成小红书风格的 HTML 预览页"""

    date_str = target_date.isoformat()
    theme_name = _safe_str(theme.get("主题名称", ""))
    theme_id = _safe_str(theme.get("theme_id", ""))

    # 小红书标题
    xhs_title = _safe_str(schedule_row.get("小红书标题")) or f"{ctx.get('目的地', '')}{theme_name}"

    # 资源卡片 HTML
    def resource_card(res: dict, label: str) -> str:
        if not res:
            return ""
        rid = _safe_str(res.get("resource_id", ""))
        rtype = _safe_str(res.get("resource_type", ""))
        name = _safe_str(res.get("product_name_cn", ""))
        area = _safe_str(res.get("area", ""))
        price = _safe_str(res.get("price_text", ""))
        highlights = _safe_str(res.get("highlights", ""))
        vk = _safe_str(res.get("visual_keywords", ""))
        angle = _safe_str(res.get("xhs_angle", ""))

        return f"""
        <div class="resource-card {label.lower()}">
            <div class="resource-label">{label}</div>
            <div class="resource-header">
                <span class="resource-id">{rid}</span>
                <span class="resource-type">{rtype}</span>
            </div>
            <h3 class="resource-name">{name}</h3>
            <div class="resource-meta">
                <span>📍 {area}</span>
                <span>💰 {price}</span>
            </div>
            {f'<p class="resource-highlights">✨ {highlights}</p>' if highlights else ''}
            {f'<p class="resource-keywords">🏷️ {vk}</p>' if vk else ''}
            {f'<p class="resource-angle">📝 {angle}</p>' if angle else ''}
        </div>"""

    # 图片 Prompt 卡片
    prompt_cards = ""
    for ip in image_prompts_filled:
        prompt_cards += f"""
        <div class="prompt-card">
            <div class="prompt-type-badge">{ip['prompt_type']}</div>
            <div class="prompt-text">{ip['filled_prompt']}</div>
            <div class="prompt-meta">
                <span>🖼️ 画幅：{ip['aspect']}</span>
                <span>🚫 限制：{ip['negative']}</span>
            </div>
        </div>"""

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>BWS 小红书内容预览 - {date_str}</title>
<style>
  :root {{
    --xhs-red: #ff2442;
    --xhs-red-light: #fff0f2;
    --bg: #f5f5f5;
    --card-bg: #ffffff;
    --text: #333333;
    --text-secondary: #999999;
    --border: #e8e8e8;
    --radius: 12px;
    --shadow: 0 2px 12px rgba(0,0,0,0.08);
  }}

  * {{ margin: 0; padding: 0; box-sizing: border-box; }}

  body {{
    font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
    background: var(--bg);
    color: var(--text);
    line-height: 1.6;
  }}

  .container {{
    max-width: 960px;
    margin: 0 auto;
    padding: 20px;
  }}

  /* Header */
  .header {{
    text-align: center;
    padding: 30px 20px;
    background: linear-gradient(135deg, var(--xhs-red), #ff6b81);
    color: white;
    border-radius: var(--radius);
    margin-bottom: 24px;
    box-shadow: var(--shadow);
  }}
  .header h1 {{ font-size: 24px; font-weight: 700; margin-bottom: 8px; }}
  .header .date {{ font-size: 14px; opacity: 0.9; }}
  .header .theme-tag {{
    display: inline-block;
    margin-top: 8px;
    padding: 4px 14px;
    background: rgba(255,255,255,0.25);
    border-radius: 20px;
    font-size: 13px;
  }}

  /* Section */
  .section {{
    background: var(--card-bg);
    border-radius: var(--radius);
    padding: 24px;
    margin-bottom: 20px;
    box-shadow: var(--shadow);
  }}
  .section-title {{
    font-size: 18px;
    font-weight: 700;
    margin-bottom: 16px;
    padding-left: 12px;
    border-left: 4px solid var(--xhs-red);
  }}

  /* Resource Cards */
  .resources-grid {{
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(280px, 1fr));
    gap: 16px;
  }}
  .resource-card {{
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 16px;
    position: relative;
    transition: box-shadow 0.2s;
  }}
  .resource-card:hover {{ box-shadow: 0 4px 16px rgba(0,0,0,0.12); }}
  .resource-label {{
    position: absolute;
    top: -10px;
    left: 12px;
    padding: 2px 12px;
    border-radius: 10px;
    font-size: 12px;
    font-weight: 600;
    color: white;
  }}
  .main .resource-label {{ background: var(--xhs-red); }}
  .mix1 .resource-label {{ background: #ff9500; }}
  .mix2 .resource-label {{ background: #34c759; }}
  .resource-header {{
    display: flex;
    gap: 8px;
    align-items: center;
    margin-top: 6px;
    margin-bottom: 8px;
  }}
  .resource-id {{ font-size: 12px; color: var(--text-secondary); }}
  .resource-type {{
    font-size: 11px;
    padding: 2px 8px;
    background: var(--xhs-red-light);
    color: var(--xhs-red);
    border-radius: 4px;
  }}
  .resource-name {{ font-size: 16px; font-weight: 600; margin-bottom: 6px; }}
  .resource-meta {{
    display: flex;
    gap: 16px;
    font-size: 13px;
    color: var(--text-secondary);
    margin-bottom: 8px;
  }}
  .resource-highlights, .resource-keywords {{
    font-size: 13px;
    color: #666;
    margin-bottom: 4px;
  }}
  .resource-angle {{
    font-size: 12px;
    color: var(--text-secondary);
    font-style: italic;
  }}

  /* Post Preview — 小红书卡片风格 */
  .xhs-phone-frame {{
    max-width: 375px;
    margin: 0 auto;
    background: white;
    border-radius: var(--radius);
    overflow: hidden;
    box-shadow: var(--shadow);
  }}
  .xhs-cover {{
    width: 100%;
    aspect-ratio: 3/4;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    display: flex;
    flex-direction: column;
    justify-content: flex-end;
    padding: 24px;
    color: white;
    position: relative;
  }}
  .xhs-cover::after {{
    content: '';
    position: absolute;
    bottom: 0;
    left: 0;
    right: 0;
    height: 50%;
    background: linear-gradient(transparent, rgba(0,0,0,0.6));
  }}
  .xhs-cover-text {{
    position: relative;
    z-index: 1;
  }}
  .xhs-cover-text h2 {{
    font-size: 20px;
    font-weight: 700;
    line-height: 1.4;
    text-shadow: 0 1px 4px rgba(0,0,0,0.3);
  }}
  .xhs-cover-text .cover-subtitle {{
    font-size: 13px;
    margin-top: 6px;
    opacity: 0.9;
  }}
  .xhs-body {{
    padding: 16px;
  }}
  .xhs-body .author {{
    display: flex;
    align-items: center;
    gap: 10px;
    margin-bottom: 12px;
  }}
  .xhs-body .avatar {{
    width: 36px;
    height: 36px;
    border-radius: 50%;
    background: linear-gradient(135deg, #ff6b81, #ff2442);
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-weight: 700;
    font-size: 14px;
  }}
  .xhs-body .author-info .name {{ font-size: 14px; font-weight: 600; }}
  .xhs-body .author-info .tag {{ font-size: 11px; color: var(--text-secondary); }}
  .xhs-body .post-content {{
    font-size: 14px;
    line-height: 1.8;
    white-space: pre-wrap;
    color: var(--text);
  }}
  .xhs-body .post-tags {{
    margin-top: 12px;
    font-size: 12px;
    color: #13386f;
  }}

  /* Prompt Cards */
  .prompt-card {{
    border: 1px dashed var(--xhs-red);
    border-radius: var(--radius);
    padding: 16px;
    margin-bottom: 12px;
    background: var(--xhs-red-light);
  }}
  .prompt-type-badge {{
    display: inline-block;
    padding: 2px 10px;
    background: var(--xhs-red);
    color: white;
    border-radius: 4px;
    font-size: 12px;
    font-weight: 600;
    margin-bottom: 8px;
  }}
  .prompt-text {{
    font-size: 13px;
    line-height: 1.6;
    color: #333;
    font-family: "SF Mono", "Consolas", "Monaco", monospace;
    word-break: break-all;
  }}
  .prompt-meta {{
    margin-top: 8px;
    font-size: 12px;
    color: var(--text-secondary);
    display: flex;
    gap: 16px;
  }}

  /* Footer */
  .footer {{
    text-align: center;
    padding: 20px;
    color: var(--text-secondary);
    font-size: 12px;
  }}
</style>
</head>
<body>
<div class="container">

  <!-- Header -->
  <div class="header">
    <h1>BWS 小红书内容预览</h1>
    <div class="date">📅 {date_str}</div>
    <div class="theme-tag">{theme_id} · {theme_name}</div>
  </div>

  <!-- 选品资源 -->
  <div class="section">
    <div class="section-title">📦 今日选品资源</div>
    <div class="resources-grid">
      {resource_card(main, "主资源")}
      {resource_card(mix1, "搭配1") if mix1 else '<div style="color:#999;font-size:13px;">无搭配资源</div>'}
      {resource_card(mix2, "搭配2") if mix2 else ''}
    </div>
  </div>

  <!-- 小红书正文预览 -->
  <div class="section">
    <div class="section-title">📝 小红书正文预览</div>
    <div class="xhs-phone-frame">
      <div class="xhs-cover">
        <div class="xhs-cover-text">
          <h2>{xhs_title}</h2>
          <div class="cover-subtitle">{ctx.get("封面卖点", "")}</div>
        </div>
      </div>
      <div class="xhs-body">
        <div class="author">
          <div class="avatar">B</div>
          <div class="author-info">
            <div class="name">BWS印尼旅行</div>
            <div class="tag">专业印尼地接 · 私信定制行程</div>
          </div>
        </div>
        <div class="post-content">{post_text}</div>
      </div>
    </div>
  </div>

  <!-- 图片 Prompt -->
  <div class="section">
    <div class="section-title">🎨 图片 Prompt（占位符，待接入 gpt-image）</div>
    {prompt_cards}
  </div>

  <!-- 原始数据 JSON -->
  <div class="section">
    <div class="section-title">🔍 上下文变量（调试用）</div>
    <pre style="font-size:12px;color:#666;overflow-x:auto;white-space:pre-wrap;">{json.dumps(ctx, ensure_ascii=False, indent=2)}</pre>
  </div>

  <div class="footer">
    BWS 小红书内容生成器 · {date_str} · 自动生成，人工审核后发布
  </div>
</div>
</body>
</html>"""
    return html


# ═══════════════════════════════════════════════════════════════════════════════
# 6. 主流程
# ═══════════════════════════════════════════════════════════════════════════════

def generate_for_date(target_date: date, data: dict) -> Path:
    """为指定日期生成完整内容包，返回输出目录路径"""
    print(f"\n{'='*60}")
    print(f"  🎯 生成内容：{target_date.isoformat()}")
    print(f"{'='*60}")

    # --- Step 1: 匹配今日选题 ---
    schedule_row = get_schedule_for_date(data["schedule"], target_date)
    if not schedule_row:
        print(f"  ⚠️ 未找到 {target_date} 的选题排期，跳过")
        return Path("")

    theme_id = _safe_str(schedule_row.get("主题ID"))
    print(f"  📋 主题：{theme_id} - {_safe_str(schedule_row.get('主题名称'))}")

    # --- Step 2: 查找主题配置 ---
    theme = None
    for t in data["themes"]:
        if _safe_str(t.get("theme_id")) == theme_id:
            theme = t
            break
    if not theme:
        print(f"  ⚠️ 未找到主题 {theme_id} 的配置，跳过")
        return Path("")

    # --- Step 3: 筛选资源 ---
    main_hint = _safe_str(schedule_row.get("主资源ID")) or None
    main = pick_main_resource(data["resources"], theme_id, hint_id=main_hint)
    print(f"  🏆 主资源：{main.get('resource_id')} - {main.get('product_name_cn')}")

    mix1_hint = _safe_str(schedule_row.get("搭配资源ID 1")) or None
    mix1 = pick_mix_resource(data["resources"], theme_id, main, slot=1)
    if mix1:
        print(f"  🔗 搭配1：{mix1.get('resource_id')} - {mix1.get('product_name_cn')}")

    mix2 = pick_mix_resource(data["resources"], theme_id, main, slot=2)
    if mix2:
        print(f"  🔗 搭配2：{mix2.get('resource_id')} - {mix2.get('product_name_cn')}")

    # --- Step 4: 构建文案上下文 + 填充模板 ---
    ctx = build_copy_context(theme, main, mix1, mix2)
    filled_modules = fill_copy_templates(data["copy_templates"], ctx)
    post_text = generate_full_post(filled_modules, ctx, theme)
    print(f"  📝 文案生成完成，{len(post_text)} 字")

    # --- Step 5: 填充图片 Prompt ---
    image_prompts_filled = fill_image_prompts(data["image_prompts"], ctx, theme_id)
    print(f"  🎨 图片 Prompt 生成完成，{len(image_prompts_filled)} 条")

    # --- Step 6: 输出文件 ---
    out_dir = OUTPUT_DIR / target_date.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)

    # JSON 数据包
    output_data = {
        "date": target_date.isoformat(),
        "theme_id": theme_id,
        "theme_name": _safe_str(theme.get("主题名称")),
        "schedule_row": {k: _safe_str(v) for k, v in schedule_row.items()},
        "main_resource": {k: _safe_str(v) for k, v in main.items()},
        "mix1_resource": {k: _safe_str(v) for k, v in mix1.items()} if mix1 else None,
        "mix2_resource": {k: _safe_str(v) for k, v in mix2.items()} if mix2 else None,
        "filled_copy_modules": filled_modules,
        "full_post_text": post_text,
        "image_prompts": image_prompts_filled,
        "context_variables": ctx,
    }

    json_path = out_dir / "content_data.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    print(f"  💾 JSON → {json_path}")

    # 纯文案 txt
    txt_path = out_dir / "post_text.txt"
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(post_text)
    print(f"  📄 TXT  → {txt_path}")

    # 图片 Prompt txt
    prompts_path = out_dir / "image_prompts.txt"
    with open(prompts_path, "w", encoding="utf-8") as f:
        for ip in image_prompts_filled:
            f.write(f"[{ip['prompt_type']}]\n")
            f.write(f"Prompt: {ip['filled_prompt']}\n")
            f.write(f"Negative: {ip['negative']}\n")
            f.write(f"Aspect: {ip['aspect']}\n\n")
    print(f"  🖼️ Prompts → {prompts_path}")

    # HTML 预览页
    html = generate_html_preview(
        target_date, schedule_row, theme, main, mix1, mix2,
        post_text, filled_modules, image_prompts_filled, ctx,
    )
    html_path = out_dir / "preview.html"
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  🌐 HTML → {html_path}")

    return out_dir


def main():
    parser = argparse.ArgumentParser(description="BWS 小红书每日内容生成器")
    parser.add_argument("date", nargs="?", help="目标日期 (YYYY-MM-DD)，默认今天")
    parser.add_argument("--date", dest="date_flag", help="目标日期 (YYYY-MM-DD)")
    parser.add_argument("--all", action="store_true", help="生成全部排期内容")
    args = parser.parse_args()

    # 解析日期
    if args.all:
        target_dates = None  # 后面从 schedule 中读取
    else:
        date_str = args.date or args.date_flag
        if date_str:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        else:
            target_date = date.today()
        target_dates = [target_date]

    # 加载数据
    print("📖 加载 Excel 数据...")
    data = load_all_data()
    print(f"  统一资源库：{len(data['resources'])} 条")
    print(f"  内容主题库：{len(data['themes'])} 个主题")
    print(f"  每日选题模板：{len(data['schedule'])} 行排期")
    print(f"  图片Prompt模板：{len(data['image_prompts'])} 种")
    print(f"  小红书文案模板：{len(data['copy_templates'])} 个模块")

    if target_dates is None:
        # --all 模式：从 schedule 中提取所有日期
        target_dates = []
        for row in data["schedule"]:
            d = _parse_date(row.get("发布日期"))
            if d and d not in target_dates:
                target_dates.append(d)
        target_dates.sort()
        print(f"\n  📅 共 {len(target_dates)} 个排期日期")

    # 生成
    output_dirs = []
    for td in target_dates:
        out_dir = generate_for_date(td, data)
        if out_dir and out_dir.exists():
            output_dirs.append(out_dir)

    # 汇总
    print(f"\n{'='*60}")
    print(f"  ✅ 完成！共生成 {len(output_dirs)} 个内容包")
    for d in output_dirs:
        print(f"     → {d}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()
