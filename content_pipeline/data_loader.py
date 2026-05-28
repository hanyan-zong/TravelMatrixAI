from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl

from .config import DEFAULT_EXCEL_PATH


def sheet_to_dicts(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[dict[str, Any]]:
    headers: list[str | None] = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        row: dict[str, Any] = {}
        for c, header in enumerate(headers, 1):
            if header is not None:
                row[str(header)] = ws.cell(r, c).value
        if any(v is not None for v in row.values()):
            rows.append(row)
    return rows


def load_all_data(excel_path: Path = DEFAULT_EXCEL_PATH) -> dict[str, list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(str(excel_path), read_only=False, data_only=True)
    try:
        return {
            "resources": sheet_to_dicts(wb["统一资源库"]),
            "themes": sheet_to_dicts(wb["内容主题库"]),
            "schedule": sheet_to_dicts(wb["每日选题模板"]),
            "image_prompts": sheet_to_dicts(wb["图片Prompt模板"]),
            "copy_templates": sheet_to_dicts(wb["小红书文案模板"]),
        }
    finally:
        wb.close()

