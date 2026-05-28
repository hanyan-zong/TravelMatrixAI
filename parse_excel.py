"""解析 BWS 旅行资源 Excel → JSON

用法: python parse_excel.py [excel_path]
"""
import json
import sys
from pathlib import Path

import openpyxl

from parsers import PARSERS

EXCEL_DEFAULT = Path(__file__).parent / "BWS碎片化整理 - 适配AI模板.xlsx"
DATA_DIR = Path(__file__).parent / "data"


def load_surcharge_notes(wb):
    """从'注意事项'sheet读取各产品的旺季附加费说明"""
    if "注意事项" not in wb.sheetnames:
        return {}

    ws = wb["注意事项"]
    d = {}

    for row in range(2, ws.max_row + 1):
        sheet_name = ws.cell(row, 1).value
        note = ws.cell(row, 2).value

        if sheet_name and note:
            d[str(sheet_name).strip()] = str(note).strip()

    return d


def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_DEFAULT
    print(f"读取: {excel_path}")

    wb = openpyxl.load_workbook(str(excel_path))
    surcharge = load_surcharge_notes(wb)

    DATA_DIR.mkdir(exist_ok=True)

    for sheet_name, parser_fn in PARSERS.items():

        if sheet_name not in wb.sheetnames:
            print(f"  ⚠ Sheet '{sheet_name}' 不存在，跳过")
            continue

        ws = wb[sheet_name]

        note = surcharge.get(sheet_name)

        if not note:

            for key, val in surcharge.items():

                if sheet_name in key or key in sheet_name:
                    note = val
                    break

        items = parser_fn(ws, surcharge_note=note)

        out_name = sheet_name.replace("+", "_").replace(" ", "_")
        out_path = DATA_DIR / f"{out_name}.json"
        data = [item.model_dump() for item in items]

        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        print(f"  ✓ {sheet_name}: {len(items)} 条 → {out_path.name}")

    all_items = []

    for json_file in sorted(DATA_DIR.glob("*.json")):

        if json_file.name == "all_resources.json":
            continue

        with open(json_file, encoding="utf-8") as f:
            all_items.extend(json.load(f))

    all_path = DATA_DIR / "all_resources.json"

    with open(all_path, "w", encoding="utf-8") as f:
        json.dump(all_items, f, ensure_ascii=False, indent=2)

    print(f"\n合计: {len(all_items)} 条资源 → {all_path.name}")
    print("完成!")


if __name__ == "__main__":
    main()
